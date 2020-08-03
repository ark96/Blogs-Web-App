import openpyxl
path = "C:\\Users\\Sangarshan Reddy\\Desktop\\CCCurrentStatement11-05-2019.xlsx"
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
total = 0
refund = 0
for i in range(1,m_row + 1):
    cell_obj1 = sheet_obj.cell(row = i, column = 2)
    cell_obj2 = sheet_obj.cell(row = i, column = 6)
    if cell_obj1.value == 'BHARAT FILLING STATION':
        if cell_obj2.value > 0:
            refund = refund + cell_obj2.value
        else:
            total = total + cell_obj2.value 
        print(cell_obj1.value,cell_obj2.value)
print("Due Amount = ",total)
print("Refund Amount = ",refund)
